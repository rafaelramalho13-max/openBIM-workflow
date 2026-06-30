import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_bep_annex():
    wb = Workbook()
    
    # Define styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    data_bold_font = Font(name=font_family, size=10, bold=True, color="333333")
    note_font = Font(name=font_family, size=9, italic=True, color="666666")
    
    fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_ice = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    fill_light_gray = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    fill_accent = PatternFill(start_color="E1EBF5", end_color="E1EBF5", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='double', color='1B365D')
    )

    # ----------------------------------------------------
    # TAB 1: Metadados do BEP
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "1. Metadados do BEP"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A1:D2")
    ws1["A1"] = "Plano de Execução BIM (BEP) - Resposta Técnica ao PIR"
    ws1["A1"].font = title_font
    ws1["A1"].fill = fill_navy
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws1.row_dimensions[1].height = 25
    ws1.row_dimensions[2].height = 25
    
    # Section Header
    ws1["A4"] = "Informações de Controle e Gestão da Informação"
    ws1["A4"].font = section_font
    ws1["A4"].alignment = align_left
    
    metadata = [
        ("Projeto:", "Demonstração de Workflow openBIM para Gestão de Informação", "Código:", "DEMO-BIM-2026"),
        ("Fase:", "Estudo de Caso / Protótipo", "Data de Emissão:", "28/06/2026"),
        ("Autor do BEP:", "Equipe de Coordenação de Projeto (Contratada)", "Status do BEP:", "Em Análise para Aprovamento"),
        ("PIR de Referência:", "PIR - Gestão de Informação v1.0", "Tipo de Modelo:", "Arquitetura (ARQ) e Estrutura (EST)")
    ]
    
    row_idx = 5
    for row in metadata:
        ws1.row_dimensions[row_idx].height = 20
        # Col A
        ws1.cell(row=row_idx, column=1, value=row[0]).font = data_bold_font
        ws1.cell(row=row_idx, column=1).fill = fill_ice
        ws1.cell(row=row_idx, column=1).alignment = align_left
        ws1.cell(row=row_idx, column=1).border = thin_border
        
        # Col B
        ws1.cell(row=row_idx, column=2, value=row[1]).font = data_font
        ws1.cell(row=row_idx, column=2).alignment = align_left
        ws1.cell(row=row_idx, column=2).border = thin_border
        
        # Col C
        ws1.cell(row=row_idx, column=3, value=row[2]).font = data_bold_font
        ws1.cell(row=row_idx, column=3).fill = fill_ice
        ws1.cell(row=row_idx, column=3).alignment = align_left
        ws1.cell(row=row_idx, column=3).border = thin_border
        
        # Col D
        ws1.cell(row=row_idx, column=4, value=row[3]).font = data_font
        ws1.cell(row=row_idx, column=4).alignment = align_left
        ws1.cell(row=row_idx, column=4).border = thin_border
        
        row_idx += 1
        
    # Team section
    ws1.cell(row=10, column=1, value="Papéis e Responsabilidades (Equipe Técnica)").font = section_font
    
    team_headers = ["Papel BIM", "Nome do Profissional", "Empresa / Organização", "Contato"]
    ws1.row_dimensions[11].height = 25
    for col_idx, text in enumerate(team_headers, 1):
        cell = ws1.cell(row=11, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    team_rows = [
        ("BIM Manager", "Mariana Silva", "BIM Solution Engineers Ltda.", "mariana.silva@bimsolution.com"),
        ("Coordenador BIM - Arq", "Carlos Eduardo", "Studio ArqDesign S.A.", "carlos.eduardo@arqdesign.com.br"),
        ("Coordenador BIM - Est", "Rodrigo Nogueira", "Engenharia Forte S/S", "rodrigo.nogueira@engforte.com.br"),
        ("Gestor de Informação (Cliente)", "Rafael Ramalho", "Contratante openBIM", "rafael.ramalho@openbim.com.br")
    ]
    
    row_idx = 12
    for row in team_rows:
        ws1.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = align_left
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = fill_light_gray
        row_idx += 1

    # ----------------------------------------------------
    # TAB 2: Resposta e Matriz de Requisitos
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="2. Matriz de Requisitos (BEP)")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Matriz de Resposta Técnica aos Requisitos de Informação de Projeto (PIR)"
    ws2["A1"].font = title_font
    ws2["A1"].fill = fill_navy
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 30
    
    ws2["A3"] = "Esta tabela descreve como a equipe de projeto responde a cada PIR através do mapeamento IFC correto."
    ws2["A3"].font = note_font
    
    matrix_headers = [
        "Cód. PIR", "Nome do Requisito", "Entidade IFC Alvo", 
        "Conjunto (Pset/Qto)", "Propriedade IFC", "Tipo de Dado IFC", "Critério BEP de Validação"
    ]
    
    ws2.row_dimensions[4].height = 28
    for col_idx, text in enumerate(matrix_headers, 1):
        cell = ws2.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    matrix_data = [
        ("PIR-01", "Espaços com LongName preenchido", "IfcSpace", "Atributo IFC", "LongName", "IfcLabel (Texto)", "Presença de IfcSpace obrigatória. Atributo LongName (Nome descritivo longo) deve ser preenchido."),
        ("PIR-02", "Volume de Pilares", "IfcColumn", "Qto_ColumnBaseQuantities", "NetVolume", "IfcVolumeMeasure (Real)", "Volume líquido deve ser calculado e > 0 na exportação IFC."),
        ("PIR-02", "Volume de Vigas", "IfcBeam", "Qto_BeamBaseQuantities", "NetVolume", "IfcVolumeMeasure (Real)", "Volume líquido deve ser calculado e > 0 na exportação IFC."),
        ("PIR-02", "Volume de Lajes", "IfcSlab", "Qto_SlabBaseQuantities", "NetVolume", "IfcVolumeMeasure (Real)", "Volume líquido deve ser calculado e > 0 na exportação IFC."),
        ("PIR-02", "Volume de Fundações", "IfcFooting", "Qto_FootingBaseQuantities", "NetVolume", "IfcVolumeMeasure (Real)", "Volume líquido deve ser calculado e > 0 na exportação IFC."),
        ("PIR-03", "Fire Rating em Paredes", "IfcWall, IfcWallStandardCase", "Pset_WallCommon", "FireRating", "IfcLabel (Texto)", "Preenchimento obrigatório para todas as paredes com classificação de incêndio (ex: '60 min').")
    ]
    
    row_idx = 5
    for row in matrix_data:
        ws2.row_dimensions[row_idx].height = 25
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [1, 3]:
                cell.alignment = align_center
                cell.font = data_bold_font
            elif col_idx in [4, 5, 6]:
                cell.alignment = align_left
                cell.fill = fill_light_gray
            else:
                cell.alignment = align_left
                
            if "PIR-01" in row[0]:
                cell.fill = fill_accent
        row_idx += 1

    # ----------------------------------------------------
    # TAB 3: Guia de Modelagem e Exportação
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="3. Guia de Exportação (Revit)")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "Guia Prático para Geração de Modelos IFC em Conformidade"
    ws3["A1"].font = title_font
    ws3["A1"].fill = fill_navy
    ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws3.row_dimensions[1].height = 30
    
    ws3["A3"] = "Instruções específicas para modeladores garantirem que os softwares exportem as propriedades esperadas no IDS."
    ws3["A3"].font = note_font
    
    guide_headers = ["Elemento", "Como Modelar no Software", "Configurações de Exportação IFC (Revit)", "Propriedade Resultante no IFC"]
    ws3.row_dimensions[4].height = 25
    for col_idx, text in enumerate(guide_headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    guide_data = [
        ("Espaços\n(IfcSpace)", 
         "Criar 'Rooms' (Ambientes) em todas as áreas internas do projeto. Preencher obrigatoriamente o parâmetro 'Long Name' (Nome Longo) nas propriedades do ambiente.",
         "No exportador IFC do Revit:\n1. Marcar 'Export rooms in IFC' para traduzir Rooms como IfcSpace.\nIsso mapeia automaticamente o 'Long Name' para o atributo LongName no IFC.",
         "Entidade: IfcSpace\nAtributo IFC: LongName"),
         
        ("Elementos Estruturais\n(Pilares, Vigas, Lajes, Fundações)",
         "Modelar elementos utilizando as ferramentas nativas de modelagem estrutural estruturadas por níveis corretos. Não utilizar 'Model In-Place' (modelagem no local) genérico.",
         "No exportador IFC do Revit:\n1. Marcar a opção 'Export base quantities' (Exportar quantidades básicas).\nIsso gerará os Property Sets de quantidades oficiais (Qto_*BaseQuantities) automaticamente contendo volumes.",
         "Entidades: IfcColumn, IfcBeam, IfcSlab, IfcFooting\nPropertySet: Qto_[Elemento]BaseQuantities\nPropriedade: NetVolume"),
         
        ("Paredes\n(IfcWall)",
         "Para cada tipo de parede que possua requisitos de incêndio, preencher o parâmetro padrão do tipo 'Fire Rating' (Classificação de Incêndio) no Revit.",
         "O Revit mapeia por padrão o parâmetro do tipo 'Fire Rating' das paredes para o Property Set regulamentar 'Pset_WallCommon' com o nome 'FireRating' na exportação oficial IFC.",
         "Entidade: IfcWall / IfcWallStandardCase\nPropertySet: Pset_WallCommon\nPropriedade: FireRating")
    ]
    
    row_idx = 5
    for row in guide_data:
        ws3.row_dimensions[row_idx].height = 70
        for col_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_center
                cell.font = data_bold_font
                cell.fill = fill_ice
            else:
                cell.alignment = align_left
        row_idx += 1

    # ----------------------------------------------------
    # Auto-adjust column widths across all sheets
    # ----------------------------------------------------
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            # We don't want to adjust width based on merged title rows
            max_len = 0
            for cell in col:
                if cell.coordinate in ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "A2", "B2", "C2", "D2", "A3", "A10"]:
                    continue
                val_str = str(cell.value or '')
                # Handle linebreaks in length calculation
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            col_letter = get_column_letter(col[0].column)
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Specific adjustment for description columns
    ws2.column_dimensions['G'].width = 45
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 40
    ws3.column_dimensions['D'].width = 35

    # Ensure requirements directory exists
    os.makedirs(os.path.dirname("/Users/rafaelramalho/Library/CloudStorage/OneDrive-Pessoal/openBIM_workflow/requisitos/Anexo_BEP.xlsx"), exist_ok=True)
    wb.save("/Users/rafaelramalho/Library/CloudStorage/OneDrive-Pessoal/openBIM_workflow/requisitos/Anexo_BEP.xlsx")
    print("Anexo_BEP.xlsx gerado com sucesso!")

if __name__ == "__main__":
    create_bep_annex()
