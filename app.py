import streamlit as st
import ifcopenshell
import pandas as pd
import xml.etree.ElementTree as ET
import plotly.graph_objects as go
import os
import glob
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set page config
st.set_page_config(
    page_title="openBIM Information Manager",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Premium Styling (CSS)
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap');
        
        html, body, [class*="css"], .stApp {
            font-family: 'Outfit', sans-serif;
        }
        
        .main-header {
            font-size: 2.2rem;
            font-weight: 700;
            color: #1B365D;
            margin-bottom: 0.5rem;
            border-bottom: 2px solid #E1EBF5;
            padding-bottom: 0.5rem;
        }
        
        .sub-header {
            font-size: 1.1rem;
            color: #555;
            margin-bottom: 2rem;
        }
        
        .metric-card {
            background-color: #F8F9FA;
            border: 1px solid #E1EBF5;
            border-radius: 10px;
            padding: 1.2rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
            text-align: center;
        }
        
        .metric-value {
            font-size: 2rem;
            font-weight: 700;
            color: #1B365D;
            margin-bottom: 0.2rem;
        }
        
        .metric-label {
            font-size: 0.9rem;
            color: #666;
            font-weight: 500;
        }
        
        .status-badge {
            padding: 4px 8px;
            border-radius: 4px;
            font-weight: bold;
            font-size: 0.85rem;
            display: inline-block;
        }
        .status-ok {
            background-color: #D4EDDA;
            color: #155724;
            border: 1px solid #C3E6CB;
        }
        .status-fail {
            background-color: #F8D7DA;
            color: #721C24;
            border: 1px solid #F5C6CB;
        }
        .status-info {
            background-color: #E2E3E5;
            color: #383D41;
            border: 1px solid #D6D8DB;
        }
    </style>
""", unsafe_allow_html=True)

# Helper functions for IFC parsing
def get_element_level(element):
    # Physical elements (Wall, Column, Beam, Slab, Footing, etc.)
    if hasattr(element, "ContainedInStructure") and element.ContainedInStructure:
        for rel in element.ContainedInStructure:
            structure = rel.RelatingStructure
            if structure.is_a("IfcBuildingStorey"):
                return structure.Name
            
    # Spatial elements (Space)
    if element.is_a("IfcSpace"):
        if hasattr(element, "Decomposes") and element.Decomposes:
            for rel in element.Decomposes:
                parent = rel.RelatingObject
                if parent.is_a("IfcBuildingStorey"):
                    return parent.Name
                
    # Fallback to decomposition
    if hasattr(element, "Decomposes") and element.Decomposes:
        for rel in element.Decomposes:
            parent = rel.RelatingObject
            if parent.is_a("IfcBuildingStorey"):
                return parent.Name
            elif parent.is_a("IfcSpace"):
                return get_element_level(parent)
                
    return "Não associado"

def get_element_reference(element):
    el_type = element.is_a()
    base_entity = el_type[3:] if el_type.startswith("Ifc") else el_type
    if "StandardCase" in base_entity:
        base_entity = base_entity.replace("StandardCase", "")
    pset_common = f"Pset_{base_entity}Common"
    has_ref, ref_val = check_property(element, pset_common, "Reference")
    if has_ref and ref_val is not None:
        return str(ref_val)
    return "Ausente"

def check_property(element, pset_name, prop_name):
    """
    Checks if a property exists in a property set or element quantity.
    Returns (exists, value_found)
    """
    for definition in getattr(element, 'IsDefinedBy', []):
        if definition.is_a('IfcRelDefinesByProperties'):
            prop_def = definition.RelatingPropertyDefinition
            if prop_def.Name == pset_name:
                if prop_def.is_a('IfcPropertySet'):
                    for prop in prop_def.HasProperties:
                        if prop.Name == prop_name:
                            if prop.is_a('IfcPropertySingleValue') and prop.NominalValue is not None:
                                return True, prop.NominalValue.wrappedValue
                            return True, None
                elif prop_def.is_a('IfcElementQuantity'):
                    for q in prop_def.Quantities:
                        if q.Name == prop_name:
                            for attr in ['VolumeValue', 'AreaValue', 'LengthValue', 'CountValue', 'WeightValue']:
                                if hasattr(q, attr):
                                    return True, getattr(q, attr)
                            return True, None
    return False, None

def find_file_in_repo(filename):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # Search in current_dir and subdirectories
    for root, dirs, files in os.walk(current_dir):
        for f in files:
            if f.lower() == filename.lower():
                return os.path.join(root, f)
    # Fallback to current working directory
    for root, dirs, files in os.walk("."):
        for f in files:
            if f.lower() == filename.lower():
                return os.path.abspath(os.path.join(root, f))
    return None

def auto_generate_bep_excel(script_path):
    import subprocess
    import sys
    if script_path and os.path.exists(script_path):
        try:
            script_dir = os.path.dirname(script_path)
            subprocess.run([sys.executable, script_path], check=True, cwd=script_dir)
            return True
        except Exception as e:
            print(f"Erro ao gerar BEP Excel automaticamente: {e}")
    return False

def generate_dynamic_bep_excel(specs):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz de Mapeamento BEP"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    
    fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Header Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Matriz de Resposta Técnica do BEP (Gerada Dinamicamente)"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 25
    
    # Section Header
    headers = ["Requisito / Regra", "Entidades IFC Alvo", "Property Set / Escopo", "Nome da Propriedade / Atributo", "Tipo de Dado", "Instruções de Modelagem/Exportação"]
    ws.row_dimensions[3].height = 25
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    row_idx = 4
    for spec in specs:
        rule_name = spec["name"]
        entities = ", ".join(spec["entities"])
        for req in spec["requirements"]:
            ws.row_dimensions[row_idx].height = 20
            ws.cell(row=row_idx, column=1, value=rule_name).font = bold_font
            ws.cell(row=row_idx, column=2, value=entities).font = normal_font
            
            pset_val = req["pset"] if req.get("type", "property") == "property" else "Atributo IFC"
            ws.cell(row=row_idx, column=3, value=pset_val).font = normal_font
            ws.cell(row=row_idx, column=4, value=req["name"]).font = bold_font
            
            # Simple data type mapping
            dt = "Texto"
            if any(x in req["name"].lower() for x in ["volume", "area", "length", "count"]):
                dt = "Real / Quantidade"
            elif "rating" in req["name"].lower() or "name" in req["name"].lower():
                dt = "Texto"
            ws.cell(row=row_idx, column=5, value=dt).font = normal_font
            
            instr = f"Garantir a presença de '{req['name']}' em '{pset_val}' com preenchimento correto."
            ws.cell(row=row_idx, column=6, value=instr).font = normal_font
            
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).border = thin_border
                
            row_idx += 1
            
    # Auto-width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.coordinate == "A1":
                continue
            max_len = max(max_len, len(str(cell.value or '')))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    wb.save(output)
    return output.getvalue()

# Parser for buildingSMART IDS 1.0 XML files
def parse_ids_xml(xml_content):
    try:
        root = ET.fromstring(xml_content)
    except Exception as e:
        st.error(f"Erro ao analisar o arquivo IDS: {e}")
        return []
    
    # Extract namespace if present
    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"
        
    specs = []
    # Find specifications
    for spec_elem in root.findall(f".//{ns}specification"):
        spec_id = spec_elem.attrib.get("id", "")
        spec_name = spec_elem.attrib.get("name", "")
        desc_elem = spec_elem.find(f"{ns}description")
        spec_desc = desc_elem.text if desc_elem is not None else ""
        
        # Applicability
        applicability_elem = spec_elem.find(f"{ns}applicability")
        entities = []
        if applicability_elem is not None:
            for entity_elem in applicability_elem.findall(f".//{ns}entity"):
                name_elem = entity_elem.find(f"{ns}name")
                if name_elem is not None:
                    sv = name_elem.find(f"{ns}simpleValue")
                    entity_name = sv.text if sv is not None else name_elem.text
                    if entity_name:
                        entities.append(entity_name.upper().strip())
                        
        # Requirements
        requirements_elem = spec_elem.find(f"{ns}requirements")
        reqs = []
        if requirements_elem is not None:
            # 1. Properties
            for prop_elem in requirements_elem.findall(f".//{ns}property"):
                pset_elem = prop_elem.find(f"{ns}propertySet")
                name_elem = prop_elem.find(f"{ns}name")
                
                if pset_elem is not None:
                    sv_pset = pset_elem.find(f"{ns}simpleValue")
                    pset_val = sv_pset.text if sv_pset is not None else pset_elem.text
                else:
                    pset_val = None
                    
                if name_elem is not None:
                    sv_name = name_elem.find(f"{ns}simpleValue")
                    name_val = sv_name.text if sv_name is not None else name_elem.text
                else:
                    name_val = None
                    
                if pset_val and name_val:
                    reqs.append({
                        "type": "property",
                        "pset": pset_val.strip(),
                        "name": name_val.strip()
                    })
            
            # 2. Attributes
            for attr_elem in requirements_elem.findall(f".//{ns}attribute"):
                name_elem = attr_elem.find(f"{ns}name")
                if name_elem is not None:
                    sv_name = name_elem.find(f"{ns}simpleValue")
                    name_val = sv_name.text if sv_name is not None else name_elem.text
                else:
                    name_val = None
                    
                if name_val:
                    reqs.append({
                        "type": "attribute",
                        "name": name_val.strip()
                    })
                    
        specs.append({
            "id": spec_id,
            "name": spec_name,
            "description": spec_desc,
            "entities": entities,
            "requirements": reqs
        })
    return specs

def run_ids_validation(ifc_file, specs):
    """
    Validates an IFC file against a parsed list of IDS specifications.
    Returns (summary_metrics, detailed_results, audited_element_ids)
    """
    detailed_results = []
    audited_element_ids = set()
    
    total_conforme = 0
    total_nao_conforme = 0
    
    for spec in specs:
        spec_id = spec["id"]
        spec_name = spec["name"]
        entities = spec["entities"]
        requirements = spec["requirements"]
        
        # Find elements matching applicability
        target_elements = []
        for ent_name in entities:
            try:
                target_elements.extend(ifc_file.by_type(ent_name))
            except:
                pass
                
        for el in target_elements:
            audited_element_ids.add(el.id())
            el_id = el.GlobalId
            el_name = getattr(el, 'Name', 'Sem Nome')
            el_type = el.is_a()
            el_level = get_element_level(el)
            
            # Check requirements
            failures = []
            prop_details = []
            
            for req in requirements:
                if req["type"] == "property":
                    pset = req["pset"]
                    prop_name = req["name"]
                    
                    exists, val = check_property(el, pset, prop_name)
                    
                    # Apply specific validation rules (PIR volume check, area check, or just presence)
                    is_valid = True
                    reason = ""
                    
                    if not exists:
                        is_valid = False
                        reason = f"Propriedade '{prop_name}' ou Pset '{pset}' ausente."
                    else:
                        # PIR-02 Specific volume rule: must be > 0
                        if prop_name in ["NetVolume", "Volume"] and (val is None or not isinstance(val, (int, float)) or val <= 0):
                            is_valid = False
                            reason = f"Volume inválido ({val}). Deve ser maior que 0."
                        # PIR-03 Specific FireRating rule: must be filled and not empty
                        elif prop_name == "FireRating" and (val is None or str(val).strip() == "" or str(val).strip() == "None" or str(val).strip() == "-"):
                            is_valid = False
                            reason = f"Fire Rating não preenchido."
                            
                    prop_details.append({
                        "pset": pset,
                        "property": prop_name,
                        "value": str(val) if val is not None else "Ausente",
                        "valid": is_valid,
                        "reason": reason
                    })
                    
                    if not is_valid:
                        failures.append(reason)
                elif req["type"] == "attribute":
                    attr_name = req["name"]
                    exists = hasattr(el, attr_name)
                    val = getattr(el, attr_name, None)
                    
                    is_valid = True
                    reason = ""
                    
                    if not exists or val is None or str(val).strip() in ["", "None", "-"]:
                        is_valid = False
                        reason = f"Atributo '{attr_name}' ausente ou vazio."
                        
                    prop_details.append({
                        "pset": "Atributo IFC",
                        "property": attr_name,
                        "value": str(val) if val is not None else "Ausente",
                        "valid": is_valid,
                        "reason": reason
                    })
                    
                    if not is_valid:
                        failures.append(reason)
            
            is_element_compliant = len(failures) == 0
            if is_element_compliant:
                total_conforme += 1
            else:
                total_nao_conforme += 1
                
            el_ref = get_element_reference(el)
            
            detailed_results.append({
                "spec_id": spec_id,
                "spec_name": spec_name,
                "element_id": el_id,
                "element_name": el_name,
                "element_type": el_type,
                "reference": el_ref,
                "level": el_level,
                "compliant": is_element_compliant,
                "details": prop_details,
                "error_reason": "; ".join(failures) if failures else "OK"
            })
            
    # Calculate not audited elements (all other IfcProducts not in audited list)
    all_products = ifc_file.by_type("IfcProduct")
    all_product_ids = {p.id() for p in all_products}
    not_audited_ids = all_product_ids - audited_element_ids
    total_not_audited = len(not_audited_ids)
    
    metrics = {
        "conforme": total_conforme,
        "nao_conforme": total_nao_conforme,
        "nao_auditado": total_not_audited,
        "total_produtos": len(all_products),
        "total_auditados": len(audited_element_ids)
    }
    
    return metrics, detailed_results, audited_element_ids

def generate_report_excel(detailed_results, metrics):
    output = BytesIO()
    wb = Workbook()
    
    # Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    
    fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Tab 1: Sumário
    ws1 = wb.active
    ws1.title = "Sumário de Auditoria"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "Relatório de Auditoria openBIM IDS"
    ws1["A1"].font = title_font
    ws1["A1"].fill = fill_navy
    ws1["A1"].alignment = align_left
    ws1.row_dimensions[1].height = 25
    
    ws1.cell(row=3, column=1, value="Métrica").font = header_font
    ws1.cell(row=3, column=1).fill = fill_navy
    ws1.cell(row=3, column=2, value="Quantidade").font = header_font
    ws1.cell(row=3, column=2).fill = fill_navy
    ws1.cell(row=3, column=2).alignment = align_center
    ws1.cell(row=3, column=3, value="Percentual").font = header_font
    ws1.cell(row=3, column=3).fill = fill_navy
    ws1.cell(row=3, column=3).alignment = align_center
    
    total = metrics["total_auditados"] or 1
    summary_data = [
        ("Elementos Conformes", metrics["conforme"], metrics["conforme"]/total, fill_green),
        ("Elementos Não Conformes", metrics["nao_conforme"], metrics["nao_conforme"]/total, fill_red),
        ("Elementos Não Auditados", metrics["nao_auditado"], metrics["nao_auditado"]/(metrics["total_produtos"] or 1), fill_gray)
    ]
    
    row_idx = 4
    for label, val, pct, fill in summary_data:
        ws1.cell(row=row_idx, column=1, value=label).font = normal_font
        ws1.cell(row=row_idx, column=1).border = thin_border
        
        c2 = ws1.cell(row=row_idx, column=2, value=val)
        c2.font = bold_font
        c2.alignment = align_center
        c2.border = thin_border
        c2.fill = fill
        
        c3 = ws1.cell(row=row_idx, column=3, value=pct)
        c3.font = normal_font
        c3.number_format = '0.0%'
        c3.alignment = align_center
        c3.border = thin_border
        
        row_idx += 1
        
    # Tab 2: Detalhes
    ws2 = wb.create_sheet(title="Detalhes da Validação")
    ws2.views.sheetView[0].showGridLines = True
    
    headers = ["Regra IDS", "ID do Elemento", "Nome", "Tipo IFC", "Referência", "Nível", "Conformidade", "Propriedade", "Valor Encontrado", "Motivo da Falha"]
    ws2.row_dimensions[1].height = 25
    for col_idx, text in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    row_idx = 2
    for r in detailed_results:
        # If there are properties details
        if r["details"]:
            for d in r["details"]:
                ws2.row_dimensions[row_idx].height = 20
                ws2.cell(row=row_idx, column=1, value=r["spec_name"]).font = normal_font
                ws2.cell(row=row_idx, column=2, value=r["element_id"]).font = bold_font
                ws2.cell(row=row_idx, column=3, value=r["element_name"]).font = normal_font
                ws2.cell(row=row_idx, column=4, value=r["element_type"]).font = normal_font
                ws2.cell(row=row_idx, column=5, value=r.get("reference", "Ausente")).font = normal_font
                ws2.cell(row=row_idx, column=6, value=r["level"]).font = normal_font
                
                c7 = ws2.cell(row=row_idx, column=7, value="Conforme" if d["valid"] else "Não Conforme")
                c7.font = bold_font
                c7.alignment = align_center
                c7.fill = fill_green if d["valid"] else fill_red
                
                ws2.cell(row=row_idx, column=8, value=f"{d['pset']}.{d['property']}").font = normal_font
                ws2.cell(row=row_idx, column=9, value=str(d["value"])).font = normal_font
                ws2.cell(row=row_idx, column=10, value=d["reason"] if not d["valid"] else "OK").font = normal_font
                
                for c in range(1, 11):
                    ws2.cell(row=row_idx, column=c).border = thin_border
                row_idx += 1
        else:
            ws2.row_dimensions[row_idx].height = 20
            ws2.cell(row=row_idx, column=1, value=r["spec_name"]).font = normal_font
            ws2.cell(row=row_idx, column=2, value=r["element_id"]).font = bold_font
            ws2.cell(row=row_idx, column=3, value=r["element_name"]).font = normal_font
            ws2.cell(row=row_idx, column=4, value=r["element_type"]).font = normal_font
            ws2.cell(row=row_idx, column=5, value=r.get("reference", "Ausente")).font = normal_font
            ws2.cell(row=row_idx, column=6, value=r["level"]).font = normal_font
            
            c7 = ws2.cell(row=row_idx, column=7, value="Conforme" if r["compliant"] else "Não Conforme")
            c7.font = bold_font
            c7.alignment = align_center
            c7.fill = fill_green if r["compliant"] else fill_red
            
            ws2.cell(row=row_idx, column=8, value="Presença").font = normal_font
            ws2.cell(row=row_idx, column=9, value="Existente").font = normal_font
            ws2.cell(row=row_idx, column=10, value=r["error_reason"]).font = normal_font
            
            for c in range(1, 11):
                ws2.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1

    # Auto-width
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.coordinate in ["A1", "B1", "C1"]:
                    continue
                max_len = max(max_len, len(str(cell.value or '')))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    wb.save(output)
    return output.getvalue()

# ----------------------------------------------------
# MAIN UI
# ----------------------------------------------------

# Header
st.markdown("<div class='main-header'>🏢 Auditoria & Validação openBIM</div>", unsafe_allow_html=True)
st.markdown("<div class='sub-header'>Validação genérica de regras IDS (buildingSMART) em modelos IFC e Demonstração de Requisitos PIR / BEP</div>", unsafe_allow_html=True)

# Session States
if "custom_ifcs" not in st.session_state:
    st.session_state.custom_ifcs = {}
if "custom_ids" not in st.session_state:
    st.session_state.custom_ids = {}

# Sidebar Configuration
with st.sidebar:
    st.image("https://img.icons8.com/color/96/structural.png", width=60)
    st.header("Painel de Controle")
    
    # 1. Select / Upload IFC
    st.subheader("1. Modelos IFC")
    local_ifc_files = sorted(glob.glob("modelos/*.ifc"))
    local_ifc_names = {os.path.basename(p): p for p in local_ifc_files}
    
    all_ifc_options = {**local_ifc_names, **st.session_state.custom_ifcs}
    
    selected_ifc_name = st.selectbox(
        "Selecione o Modelo IFC:",
        options=list(all_ifc_options.keys()) if all_ifc_options else ["Nenhum modelo disponível"],
        help="Modelos presentes na pasta 'modelos/' ou carregados manualmente."
    )
    
    uploaded_ifc = st.file_uploader("Upload de novo IFC:", type=["ifc"])
    if uploaded_ifc is not None:
        filename = uploaded_ifc.name
        if filename not in st.session_state.custom_ifcs:
            # Save temporary inside workspace
            os.makedirs("modelos/temp", exist_ok=True)
            temp_path = os.path.join("modelos/temp", filename)
            with open(temp_path, "wb") as f:
                f.write(uploaded_ifc.getbuffer())
            st.session_state.custom_ifcs[filename] = temp_path
            st.success(f"IFC '{filename}' carregado com sucesso!")
            st.rerun()

    st.markdown("---")

    # 2. Select / Upload IDS
    st.subheader("2. Regras IDS")
    local_ids_files = sorted(glob.glob("requisitos/*.ids"))
    local_ids_names = {os.path.basename(p): p for p in local_ids_files}
    
    all_ids_options = {**local_ids_names, **st.session_state.custom_ids}
    
    selected_ids_name = st.selectbox(
        "Selecione as Regras IDS:",
        options=list(all_ids_options.keys()) if all_ids_options else ["Nenhuma regra disponível"],
        help="Arquivos IDS (.ids) na pasta 'requisitos/' ou carregados manualmente."
    )
    
    uploaded_ids = st.file_uploader("Upload de novo IDS:", type=["ids", "xml"])
    if uploaded_ids is not None:
        filename = uploaded_ids.name
        if filename not in st.session_state.custom_ids:
            os.makedirs("requisitos/temp", exist_ok=True)
            temp_path = os.path.join("requisitos/temp", filename)
            with open(temp_path, "wb") as f:
                f.write(uploaded_ids.getbuffer())
            st.session_state.custom_ids[filename] = temp_path
            st.success(f"IDS '{filename}' carregado com sucesso!")
            st.rerun()
            
    st.markdown("---")
    
    # 3. Action button
    btn_validate = st.button("🚀 Executar Validação openBIM", type="primary", use_container_width=True)

# ----------------------------------------------------
# VALIDATION PROCESSING
# ----------------------------------------------------
validation_run = False
if btn_validate and selected_ifc_name in all_ifc_options and selected_ids_name in all_ids_options:
    ifc_path = all_ifc_options[selected_ifc_name]
    ids_path = all_ids_options[selected_ids_name]
    
    with st.spinner("Lendo modelo IFC e regras IDS..."):
        # Load IDS
        try:
            with open(ids_path, "r", encoding="utf-8") as f:
                ids_content = f.read()
            specs = parse_ids_xml(ids_content)
        except Exception as e:
            st.error(f"Erro ao ler arquivo IDS: {e}")
            specs = []
            
        # Load IFC
        try:
            ifc_file = ifcopenshell.open(ifc_path)
            # Store elements counts
            schema = ifc_file.schema
        except Exception as e:
            st.error(f"Erro ao abrir arquivo IFC: {e}")
            ifc_file = None
            
    if ifc_file and specs:
        with st.spinner("Executando motor de validação IDS..."):
            metrics, detailed_results, audited_ids = run_ids_validation(ifc_file, specs)
            
            # Save results to session state
            st.session_state.last_validation = {
                "ifc_name": selected_ifc_name,
                "ids_name": selected_ids_name,
                "schema": schema,
                "metrics": metrics,
                "detailed_results": detailed_results,
                "specs": specs,
                "ifc_file": ifc_file  # Keep ref in session state
            }
            st.success("Validação concluída com sucesso!")
            validation_run = True

# Retrieve validation results if they exist in session state
validation_data = st.session_state.get("last_validation", None)

# Tabs definitions
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Dashboard Geral", 
    "🎯 Validação Detalhada (IDS)", 
    "📋 Demonstração PIR (Tabelas)",
    "📑 Resposta Técnica (BEP)", 
    "💾 Exportar Auditoria"
])

# ----------------------------------------------------
# TAB 1: DASHBOARD GERAL
# ----------------------------------------------------
with tab1:
    if validation_data:
        st.markdown(f"### Resultados da Validação do Modelo `{validation_data['ifc_name']}` contra `{validation_data['ids_name']}`")
        
        metrics = validation_data["metrics"]
        
        # Metrics row
        col1, col2, col3, col4 = st.columns(4)
        
        # Compliance rate
        total_audited = metrics["total_auditados"]
        compliance_rate = (metrics["conforme"] / total_audited * 100) if total_audited > 0 else 0.0
        
        with col1:
            st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value'>{metrics['total_produtos']}</div>
                    <div class='metric-label'>Total de Elementos IFC (IfcProduct)</div>
                </div>
            """, unsafe_allow_html=True)
            
        with col2:
            st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value'>{metrics['total_auditados']}</div>
                    <div class='metric-label'>Elementos Auditados pelo IDS</div>
                </div>
            """, unsafe_allow_html=True)
            
        with col3:
            st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value'>{metrics['conforme']} / {metrics['nao_conforme']}</div>
                    <div class='metric-label'>Conformes / Não Conformes</div>
                </div>
            """, unsafe_allow_html=True)
            
        with col4:
            st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value'>{compliance_rate:.1f}%</div>
                    <div class='metric-label'>Taxa de Conformidade dos Auditados</div>
                </div>
            """, unsafe_allow_html=True)
            
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Grid layout for Chart and details
        chart_col, info_col = st.columns([2, 3])
        
        with chart_col:
            # Doughnut chart
            labels = ['Conforme', 'Não Conforme', 'Não Auditado']
            values = [metrics['conforme'], metrics['nao_conforme'], metrics['nao_auditado']]
            colors = ['#28A745', '#DC3545', '#6C757D']
            
            fig = go.Figure(data=[go.Pie(
                labels=labels, 
                values=values, 
                hole=.4,
                marker_colors=colors,
                textinfo='value+percent',
                insidetextorientation='horizontal'
            )])
            
            fig.update_layout(
                title_text="Distribuição de Elementos do Modelo",
                annotations=[dict(text='Status', x=0.5, y=0.5, font_size=20, showarrow=False)],
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.1, xanchor="center", x=0.5),
                margin=dict(t=40, b=40, l=0, r=0)
            )
            st.plotly_chart(fig, use_container_width=True)
            
        with info_col:
            st.markdown("### Resumo do Diagnóstico")
            
            if compliance_rate == 100:
                st.balloons()
                st.success("🏆 Excelente! O modelo atende a 100% dos requisitos de informação mapeados no IDS.")
            elif compliance_rate >= 70:
                st.warning("⚠️ Bom nível de conformidade, mas existem dados cruciais ausentes que precisam ser resolvidos antes da emissão final.")
            else:
                st.error("🚨 Alerta! Alto índice de não conformidade encontrado. O modelo precisa ser reavaliado pela equipe de modelagem em conformidade com as orientações do BEP.")
                
            # Quick summary table by IFC entity
            detailed_results = validation_data["detailed_results"]
            df_det = pd.DataFrame(detailed_results)
            if not df_det.empty:
                st.markdown("**Resumo de Auditoria por Tipo de Elemento (IFC Entity):**")
                df_summary = df_det.groupby(["element_type", "compliant"]).size().unstack(fill_value=0)
                # Ensure columns exist
                for col in [True, False]:
                    if col not in df_summary.columns:
                        df_summary[col] = 0
                df_summary = df_summary.rename(columns={True: "Conforme", False: "Não Conforme"})
                df_summary["Total Auditado"] = df_summary["Conforme"] + df_summary["Não Conforme"]
                df_summary["Conformidade %"] = (df_summary["Conforme"] / df_summary["Total Auditado"] * 100).round(1).astype(str) + "%"
                st.dataframe(df_summary, use_container_width=True)
    else:
        st.markdown("### 🏢 Protótipo de Auditoria & Validação openBIM")
        st.markdown(
            "Este painel permite auditar e validar a conformidade de modelos de informação de construção (**IFC**) "
            "com base em requisitos contratuais computáveis descritos no formato **IDS**."
        )
        
        # Parse the currently selected IDS file in the sidebar to show its requirements dynamically
        preview_specs = []
        if selected_ids_name in all_ids_options:
            try:
                with open(all_ids_options[selected_ids_name], "r", encoding="utf-8") as f:
                    preview_content = f.read()
                preview_specs = parse_ids_xml(preview_content)
            except:
                pass
        
        col_guide1, col_guide2 = st.columns(2)
        with col_guide1:
            st.markdown("#### 🚀 Como Testar a Plataforma:")
            st.markdown(
                "1. **Selecione o Modelo IFC** no painel esquerdo.\n"
                "2. **Selecione as Regras IDS** no painel esquerdo.\n"
                "3. Clique em **🚀 Executar Validação openBIM**.\n"
                "4. Explore as abas de resultados e baixe os relatórios."
            )
        with col_guide2:
            st.markdown(f"#### 📋 Regras no Arquivo `{selected_ids_name}`:")
            if preview_specs:
                for spec in preview_specs:
                    req_labels = []
                    for req in spec["requirements"]:
                        if req.get("type", "property") == "property":
                            req_labels.append(f"`{req['pset']}.{req['name']}`")
                        elif req.get("type") == "attribute":
                            req_labels.append(f"Atributo `{req['name']}`")
                    entities_str = ", ".join(spec["entities"])
                    st.markdown(f"* **{spec['name']}** ({entities_str}): {', '.join(req_labels)}")
            else:
                st.info("Nenhuma especificação identificada no IDS selecionado.")
            
        # Show local models ready to audit
        st.markdown("---")
        st.markdown("### 📁 Modelos IFC Disponíveis no Servidor:")
        local_files_list = []
        for name, path in local_ifc_names.items():
            try:
                sz = os.path.getsize(path) / (1024 * 1024)
                local_files_list.append({"Nome do Arquivo": name, "Tamanho": f"{sz:.1f} MB"})
            except:
                pass
        if local_files_list:
            st.dataframe(pd.DataFrame(local_files_list), use_container_width=True)
        else:
            st.warning("Nenhum arquivo IFC padrão localizado na pasta 'modelos/'. Carregue novos arquivos IFC usando o painel lateral.")

# ----------------------------------------------------
# TAB 2: DETALHADO POR REGRA IDS
# ----------------------------------------------------
with tab2:
    if validation_data:
        st.markdown("### Detalhamento das Regras da Especificação IDS")
        specs = validation_data["specs"]
        detailed_results = validation_data["detailed_results"]
        
        df_results = pd.DataFrame(detailed_results)
        
        for spec in specs:
            spec_id = spec["id"]
            spec_name = spec["name"]
            
            with st.expander(f"📌 {spec_name} (ID: {spec_id})", expanded=True):
                st.markdown(f"**Descrição:** {spec['description']}")
                st.markdown(f"**Entidades Alvos:** {', '.join(spec['entities'])}")
                
                # Requirements description
                st.markdown("**Propriedades Requeridas:**")
                for req in spec["requirements"]:
                    if req.get("type", "property") == "property":
                        st.write(f"- Property Set: `{req['pset']}` | Propriedade: `{req['name']}`")
                    elif req.get("type") == "attribute":
                        st.write(f"- Atributo IFC: `{req['name']}`")
                
                # Filter results for this spec
                df_spec_res = df_results[df_results["spec_id"] == spec_id]
                
                if not df_spec_res.empty:
                    # Flatten property details for listing
                    rows = []
                    for idx, r in df_spec_res.iterrows():
                        if r["details"]:
                            for d in r["details"]:
                                rows.append({
                                    "ID Elemento (GlobalId)": r["element_id"],
                                    "Nome do Elemento": r["element_name"],
                                    "Classe IFC": r["element_type"],
                                    "Referência": r.get("reference", "Ausente"),
                                    "Nível": r["level"],
                                    "Pset / Qto": d["pset"],
                                    "Propriedade": d["property"],
                                    "Valor no Modelo": d["value"],
                                    "Conformidade": "✅ Conforme" if d["valid"] else "❌ Não Conforme",
                                    "Nota de Erro": d["reason"] if not d["valid"] else "OK"
                                })
                        else:
                            rows.append({
                                "ID Elemento (GlobalId)": r["element_id"],
                                "Nome do Elemento": r["element_name"],
                                "Classe IFC": r["element_type"],
                                "Referência": r.get("reference", "Ausente"),
                                "Nível": r["level"],
                                "Pset / Qto": "N/A",
                                "Propriedade": "Presença",
                                "Valor no Modelo": "Existente",
                                "Conformidade": "✅ Conforme" if r["compliant"] else "❌ Não Conforme",
                                "Nota de Erro": r["error_reason"]
                            })
                    
                    df_flat = pd.DataFrame(rows)
                    st.dataframe(
                        df_flat,
                        use_container_width=True,
                        column_config={
                            "Conformidade": st.column_config.TextColumn(
                                "Conformidade",
                                help="Status de conformidade do elemento"
                            )
                        }
                    )
                else:
                    st.info("Nenhum elemento deste tipo foi encontrado no modelo analisado.")
    else:
        st.info("Execute a validação no painel lateral para visualizar os dados detalhados por regra.")

# ----------------------------------------------------
# TAB 3: DEMONSTRAÇÃO PIR (TABELAS DE PROJETO)
# ----------------------------------------------------
with tab3:
    if validation_data:
        ifc_file = validation_data["ifc_file"]
        
        st.markdown("### Tabelas de Resultados dos Requisitos de Informação de Projeto (PIR)")
        
        # 1. PIR-01: Spaces Table
        st.markdown("#### PIR-01: Compartimentação e Cadastro de Espaços (IfcSpace)")
        spaces = ifc_file.by_type("IfcSpace")
        if spaces:
            space_rows = []
            for s in spaces:
                el_id = s.GlobalId
                name = s.Name
                long_name = getattr(s, 'LongName', 'N/A')
                level = get_element_level(s)
                
                # Check for Area (informative)
                has_area, area_val = check_property(s, "Qto_SpaceBaseQuantities", "GrossArea")
                if not has_area:
                    has_area, area_val = check_property(s, "Qto_SpaceBaseQuantities", "NetArea")
                area_str = f"{area_val:.2f} m²" if has_area and area_val is not None else "Ausente"
                
                # Check for LongName (main PIR-01 rule)
                has_longname = long_name is not None and str(long_name).strip() not in ["", "None", "-", "N/A"]
                status = "✅ Atende" if has_longname else "❌ Não atende"
                
                # Check for Reference (Pset_SpaceCommon -> Reference)
                ref_str = get_element_reference(s)
                
                space_rows.append({
                    "ID do Espaço (GlobalId)": el_id,
                    "Nome": name,
                    "LongName (Ambiente)": long_name,
                    "Referência": ref_str,
                    "Área dos Espaços": area_str,
                    "Nível (andar)": level,
                    "Status": status
                })
            st.dataframe(pd.DataFrame(space_rows), use_container_width=True)
        else:
            st.warning("Nenhum elemento IfcSpace encontrado neste modelo.")
            
        st.markdown("---")
        
        # 2. PIR-02: Volumes Table
        st.markdown("#### PIR-02: Volume de Elementos Estruturais (Pilares, Vigas, Lajes e Fundações)")
        structural_entities = ["IfcColumn", "IfcBeam", "IfcSlab", "IfcFooting"]
        struct_elements = []
        for ent in structural_entities:
            try:
                struct_elements.extend(ifc_file.by_type(ent))
            except:
                pass
                
        if struct_elements:
            struct_rows = []
            for el in struct_elements:
                el_id = el.GlobalId
                name = getattr(el, 'Name', 'Sem Nome')
                el_type = el.is_a()
                level = get_element_level(el)
                
                # Search for NetVolume in its standard quantity set
                pset_qto = f"Qto_{el_type[3:]}BaseQuantities" # strip 'Ifc' to get element name
                has_vol, vol_val = check_property(el, pset_qto, "NetVolume")
                
                # Fallbacks in case software exported under generic name
                if not has_vol:
                    has_vol, vol_val = check_property(el, "Qto_ConcreteBaseQuantities", "NetVolume")
                if not has_vol:
                    has_vol, vol_val = check_property(el, "Qto_ConcreteQuantities", "Volume")
                
                # Search for Reference in standard PsetCommon
                pset_common = f"Pset_{el_type[3:]}Common"
                has_ref, ref_val = check_property(el, pset_common, "Reference")
                ref_str = str(ref_val) if has_ref and ref_val is not None else "Ausente"
                
                status = "❌ Não atende"
                vol_str = "Ausente"
                
                if has_vol and vol_val is not None:
                    try:
                        vol_f = float(vol_val)
                        vol_str = f"{vol_f:.3f} m³"
                        if vol_f > 0:
                            status = "✅ Atende"
                    except:
                        pass
                
                struct_rows.append({
                    "ID do Objeto (GlobalId)": el_id,
                    "Nome": name,
                    "Classe IFC": el_type,
                    "Referência": ref_str,
                    "Volume": vol_str,
                    "Nível (andar)": level,
                    "Status": status
                })
            st.dataframe(pd.DataFrame(struct_rows), use_container_width=True)
        else:
            st.warning("Nenhum elemento de pilar, viga, laje ou fundação estrutural encontrado neste modelo.")
            
        st.markdown("---")
        
        # 3. PIR-03: Walls Fire Rating Table
        st.markdown("#### PIR-03: Classificação de Resistência ao Fogo em Paredes (IfcWall)")
        walls = ifc_file.by_type("IfcWall")
        if walls:
            wall_rows = []
            for w in walls:
                el_id = w.GlobalId
                name = w.Name
                level = get_element_level(w)
                
                has_fr, fr_val = check_property(w, "Pset_WallCommon", "FireRating")
                
                status = "❌ Não atende"
                fr_str = "Ausente"
                
                if has_fr and fr_val is not None and str(fr_val).strip() not in ["", "None", "-"]:
                    fr_str = str(fr_val)
                    status = "✅ Atende"
                    
                ref_str = get_element_reference(w)
                
                wall_rows.append({
                    "ID do Objeto (GlobalId)": el_id,
                    "Nome": name,
                    "Referência": ref_str,
                    "Nível (andar)": level,
                    "Fire Rating": fr_str,
                    "Status": status
                })
            st.dataframe(pd.DataFrame(wall_rows), use_container_width=True)
        else:
            st.warning("Nenhum elemento de parede (IfcWall) encontrado neste modelo.")
    else:
        st.info("Execute a validação no painel lateral para gerar as tabelas do PIR do projeto.")

# ----------------------------------------------------
# TAB 4: PLANO DE EXECUÇÃO (BEP DOCUMENTATION)
# ----------------------------------------------------
with tab4:
    st.markdown("### Plano de Execução BIM (BEP) — Resposta Técnica")
    st.markdown("O BEP é o documento que responde formalmente às exigências especificadas no PIR. Abaixo você pode visualizar a lógica técnica e fazer o download da planilha oficial de mapeamento de dados.")
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. Check if we are on the default project IDS
    is_default_project = (selected_ids_name == "requisitos.ids")
    
    if is_default_project:
        # Resolve paths using recursive search in repo
        pir_path = find_file_in_repo("PIR.md")
        script_path = find_file_in_repo("gerar_anexo_bep.py")
        
        # Set Excel path in the same folder as the generator script (or fallback)
        if script_path:
            bep_excel_path = os.path.join(os.path.dirname(script_path), "Anexo_BEP.xlsx")
        else:
            bep_excel_path = os.path.join(current_dir, "requisitos", "Anexo_BEP.xlsx")
            
        # Auto-generate if missing
        if bep_excel_path and not os.path.exists(bep_excel_path) and script_path:
            with st.spinner("Gerando arquivo Excel do BEP automaticamente..."):
                auto_generate_bep_excel(script_path)
                
        # Download excel BEP button
        if bep_excel_path and os.path.exists(bep_excel_path):
            with open(bep_excel_path, "rb") as f:
                st.download_button(
                    label="📥 Baixar Anexo de Requisitos do BEP (Excel)",
                    data=f.read(),
                    file_name="Anexo_BEP_Resposta_PIR.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Planilha Excel estruturando a matriz de resposta técnica de conformidade openBIM"
                )
        else:
            st.warning("O arquivo Anexo_BEP.xlsx não pôde ser gerado automaticamente. Verifique se o script `gerar_anexo_bep.py` foi enviado para o GitHub.")
    else:
        # Custom IDS selected: Generate BEP Excel template dynamically in memory!
        custom_specs = []
        if selected_ids_name in all_ids_options:
            try:
                with open(all_ids_options[selected_ids_name], "r", encoding="utf-8") as f:
                    custom_content = f.read()
                custom_specs = parse_ids_xml(custom_content)
            except Exception as e:
                st.error(f"Erro ao analisar o arquivo IDS customizado: {e}")
                
        if custom_specs:
            try:
                dynamic_excel_bytes = generate_dynamic_bep_excel(custom_specs)
                st.download_button(
                    label=f"📥 Baixar Matriz BEP Dinâmica para {selected_ids_name} (Excel)",
                    data=dynamic_excel_bytes,
                    file_name=f"Matriz_BEP_{selected_ids_name.split('.')[0]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Planilha Excel de resposta técnica gerada dinamicamente com base nas regras do IDS carregado."
                )
            except Exception as e:
                st.error(f"Erro ao gerar a planilha Excel dinâmica: {e}")
        else:
            st.info("Nenhuma especificação identificada no IDS customizado para gerar o modelo Excel.")
            
    st.markdown("---")
    
    # 2. Render Document/Specifications
    if is_default_project:
        # Embed and render PIR.md inside Streamlit
        if pir_path and os.path.exists(pir_path):
            with open(pir_path, "r", encoding="utf-8") as f:
                pir_markdown = f.read()
            st.markdown("### Documento Integrado: Requisitos de Informação de Projeto (PIR)")
            st.markdown(pir_markdown)
        else:
            st.info("O documento PIR.md não foi localizado no workspace do repositório.")
    else:
        # Custom IDS: Generate human-friendly requirements specification page dynamically
        st.markdown(f"### 📋 Requisitos de Qualidade da Informação (Gerados a partir de `{selected_ids_name}`)")
        st.markdown(
            "Abaixo estão os requisitos de qualidade de dados extraídos automaticamente da especificação IDS "
            "carregada no sistema:"
        )
        if custom_specs:
            for spec in custom_specs:
                st.markdown(f"#### {spec['name']}")
                st.markdown(f"**Objetivo/Descrição**: {spec['description']}")
                st.markdown(f"**Elementos Alvo**: `{', '.join(spec['entities'])}`")
                
                st.markdown("**Critérios de Aceitação**:")
                for req in spec["requirements"]:
                    if req.get("type", "property") == "property":
                        st.markdown(f"- Deve possuir a propriedade **`{req['name']}`** dentro do Property Set **`{req['pset']}`**.")
                    elif req.get("type") == "attribute":
                        st.markdown(f"- Deve possuir o atributo de entidade **`{req['name']}`** preenchido.")
                st.markdown("---")
        else:
            st.info("Nenhuma especificação encontrada no arquivo IDS para detalhamento.")

# ----------------------------------------------------
# TAB 5: EXPORTAR AUDITORIA
# ----------------------------------------------------
with tab5:
    if validation_data:
        st.markdown("### Exportar Resultados da Auditoria")
        st.markdown("Gere um relatório detalhado em formato Excel de todos os elementos auditados e seus status de conformidade para compartilhar com as equipes de projeto.")
        
        detailed_results = validation_data["detailed_results"]
        metrics = validation_data["metrics"]
        
        excel_bytes = generate_report_excel(detailed_results, metrics)
        
        st.download_button(
            label="📥 Baixar Relatório de Auditoria IDS (Excel)",
            data=excel_bytes,
            file_name=f"Relatorio_Auditoria_{validation_data['ifc_name'].split('.')[0]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    else:
        st.info("Para exportar relatórios de auditoria, você precisa carregar os dados e executar a validação no painel lateral.")
